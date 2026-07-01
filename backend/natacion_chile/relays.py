from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime
from itertools import combinations, permutations
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Iterable
import re
import unicodedata

from openpyxl import load_workbook

STROKES = ("backstroke", "breaststroke", "butterfly", "freestyle")
STROKE_LABELS = {
    "backstroke": "Espalda",
    "breaststroke": "Pecho",
    "butterfly": "Mariposa",
    "freestyle": "Crol",
}
STROKE_COLUMNS = {
    "backstroke": "50 Espalda",
    "breaststroke": "50 Pecho",
    "butterfly": "50 Mariposa",
    "freestyle": "50 Crol",
}
RELAY_DISTANCES = (50, 100)
RELAY_STYLES = {
    "medley": {
        "label": "combinado",
        "slots": (
            ("backstroke", "Espalda"),
            ("breaststroke", "Pecho"),
            ("butterfly", "Mariposa"),
            ("freestyle", "Crol"),
        ),
    },
    "freestyle": {
        "label": "libre",
        "slots": (
            ("freestyle", "Libre 1"),
            ("freestyle", "Libre 2"),
            ("freestyle", "Libre 3"),
            ("freestyle", "Libre 4"),
        ),
    },
}
RELAY_GENDER_RULES = {
    "mixed": {"label": "mixto", "gender_rule": "mixed_2f_2m"},
    "women": {"label": "mujeres", "gender_rule": "women"},
    "men": {"label": "hombres", "gender_rule": "men"},
}


def build_relay_types() -> dict[str, dict[str, object]]:
    relay_types: dict[str, dict[str, object]] = {}
    for distance in RELAY_DISTANCES:
        for style_key, style in RELAY_STYLES.items():
            for gender_key, gender in RELAY_GENDER_RULES.items():
                key = f"4x{distance}_{style_key}_{gender_key}"
                relay_types[key] = {
                    "label": f"4x{distance} {style['label']} {gender['label']}",
                    "distance_m": distance,
                    "style": style_key,
                    "gender_rule": gender["gender_rule"],
                    "slots": style["slots"],
                }
    return relay_types


RELAY_TYPES = build_relay_types()
RELAY_CATEGORIES = (
    ("premaster", "Premaster", 80, 96),
    ("100-119", "100 - 119", 100, 119),
    ("120-159", "120 - 159", 120, 159),
    ("160-199", "160 - 199", 160, 199),
    ("200-239", "200 - 239", 200, 239),
    ("240-279", "240 - 279", 240, 279),
    ("280-319", "280 - 319", 280, 319),
)
COMPETITION_YEAR = 2026


@dataclass(frozen=True)
class RelayTime:
    ms: int | None
    source: str
    text: str | None = None
    athlete_core_id: int | None = None
    competition_name: str | None = None
    competition_date: str | None = None


@dataclass(frozen=True)
class RelayAthlete:
    id: str
    full_name: str
    normalized_name: str
    gender: str
    birth_date: date | None
    birth_year: int | None
    age: int | None
    rut: str | None
    core_athlete_id: int | None
    times: dict[str, RelayTime]


@dataclass(frozen=True)
class RelaySlot:
    key: str
    label: str
    leg_order: int
    stroke: str
    stroke_label: str


@dataclass(frozen=True)
class RelayLeg:
    slot_key: str
    slot_label: str
    leg_order: int
    stroke: str
    stroke_label: str
    athlete_id: str | None
    athlete_name: str | None
    gender: str | None
    age: int | None
    time_ms: int | None
    time_text: str | None
    time_source: str | None


@dataclass(frozen=True)
class RelayValidation:
    is_valid: bool
    category_key: str | None
    category_label: str | None
    age_sum: int | None
    total_time_ms: int | None
    total_time_text: str | None
    errors: list[str]
    warnings: list[str]


@dataclass(frozen=True)
class RelayLineup:
    id: str
    relay_type: str
    category_key: str | None
    category_label: str | None
    age_sum: int | None
    total_time_ms: int | None
    total_time_text: str | None
    legs: list[RelayLeg]
    validation: RelayValidation


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_name(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    without_accents = "".join(char for char in decomposed if not unicodedata.combining(char))
    return " ".join(re.sub(r"[^a-zA-Z0-9 ]+", " ", without_accents).lower().split())


def normalize_name_match_key(value: str) -> str:
    return " ".join(sorted(normalize_name(value).split()))


def normalize_rut(value: object) -> str | None:
    text = str(value or "").upper()
    normalized = re.sub(r"[^0-9K]", "", text)
    return normalized or None


def parse_seed_time(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return ((value.minute * 60) + value.second) * 1000 + int(value.microsecond / 1000)
    text = str(value).strip()
    if not text:
        return None
    match = re.fullmatch(r"(?:(\d{1,2}):)?(\d{1,2})(?:[\.,](\d{1,2}))?", text)
    if not match:
        raise ValueError(f"Tiempo inválido: {text}")
    minutes = int(match.group(1) or 0)
    seconds = int(match.group(2))
    hundredths = (match.group(3) or "0").ljust(2, "0")[:2]
    return ((minutes * 60) + seconds) * 1000 + int(hundredths) * 10


def format_time(ms: int | None) -> str | None:
    if ms is None:
        return None
    minutes, remainder = divmod(ms, 60_000)
    seconds, millis = divmod(remainder, 1000)
    return f"{minutes:02d}:{seconds:02d}.{millis // 10:02d}"


def normalize_gender(value: object) -> str | None:
    text = str(value or "").strip().lower()
    if text in {"femenino", "female", "f", "mujer"}:
        return "female"
    if text in {"masculino", "male", "m", "hombre", "varon", "varón"}:
        return "male"
    return None


def compute_age(birth_date: date | datetime | None, competition_year: int = COMPETITION_YEAR) -> int | None:
    if birth_date is None:
        return None
    return competition_year - birth_date.year


def relay_category(age_sum: int | None) -> tuple[str, str] | tuple[None, None]:
    if age_sum is None:
        return None, None
    for key, label, minimum, maximum in RELAY_CATEGORIES:
        if minimum <= age_sum <= maximum:
            return key, label
    return None, None


def get_relay_config(relay_type: str) -> dict[str, object]:
    if relay_type not in RELAY_TYPES:
        raise ValueError(f"Tipo de relevo no soportado: {relay_type}")
    return RELAY_TYPES[relay_type]


def relay_distance_m(relay_type: str) -> int:
    return int(get_relay_config(relay_type)["distance_m"])


def relay_slots(relay_type: str) -> list[RelaySlot]:
    config = get_relay_config(relay_type)
    return [
        RelaySlot(key=f"leg_{index}", label=label, leg_order=index, stroke=stroke, stroke_label=STROKE_LABELS[stroke])  # type: ignore[call-arg]
        for index, (stroke, label) in enumerate(config["slots"], start=1)  # type: ignore[index]
    ]


def row_cell(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def make_excel_athlete(
    row_number: int,
    first_name: object,
    paternal: object,
    maternal: object,
    rut_value: object,
    birth_value: object,
    gender_value: object,
    times: dict[str, RelayTime] | None = None,
    competition_year: int = COMPETITION_YEAR,
) -> RelayAthlete | None:
    first_name_text = str(first_name or "").strip()
    paternal_text = str(paternal or "").strip()
    maternal_text = str(maternal or "").strip()
    if not first_name_text and not paternal_text:
        return None
    full_name = " ".join(part for part in [first_name_text, paternal_text, maternal_text] if part) or f"Fila {row_number}"
    birth_date = birth_value.date() if isinstance(birth_value, datetime) else birth_value if isinstance(birth_value, date) else None
    gender = normalize_gender(gender_value) or "unknown"
    rut = normalize_rut(rut_value)
    return RelayAthlete(
        id=rut or f"row-{row_number}",
        full_name=full_name,
        normalized_name=normalize_name(full_name),
        gender=gender,
        birth_date=birth_date,
        birth_year=birth_date.year if birth_date else None,
        age=compute_age(birth_date, competition_year),
        rut=rut,
        core_athlete_id=None,
        times=times or {stroke: RelayTime(ms=None, source="missing") for stroke in STROKES},
    )


def parse_forms_style_sheet(rows: list[tuple[object, ...]], competition_year: int) -> list[RelayAthlete] | None:
    if not rows:
        return []
    headers = [normalize_header(value) for value in rows[0]]
    header_index = {normalize_name(header): index for index, header in enumerate(headers)}
    aliases = {
        "first_name": "primer nombre",
        "paternal": "apellido paterno",
        "maternal": "apellido materno",
        "rut": "rut",
        "birth_date": "fecha de nacimiento",
        "gender": "genero",
        "breaststroke": "50 pecho",
        "backstroke": "50 espalda",
        "butterfly": "50 mariposa",
        "freestyle": "50 crol",
    }
    required = ["first_name", "paternal", "birth_date", "gender"]
    if any(aliases[key] not in header_index for key in required):
        return None

    athletes: list[RelayAthlete] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None for value in row):
            continue

        def cell(key: str) -> object:
            return row_cell(row, header_index.get(aliases[key]))

        excel_times = {
            stroke: RelayTime(ms=parse_seed_time(cell(stroke)), source="excel")
            for stroke in STROKES
        }
        athlete = make_excel_athlete(
            row_number=row_number,
            first_name=cell("first_name"),
            paternal=cell("paternal"),
            maternal=cell("maternal"),
            rut_value=cell("rut"),
            birth_value=cell("birth_date"),
            gender_value=cell("gender"),
            times=excel_times,
            competition_year=competition_year,
        )
        if athlete:
            athletes.append(athlete)
    return athletes


def stroke_from_event_label(value: object) -> str | None:
    label = normalize_name(str(value or ""))
    if "50" not in label:
        return None
    if "espalda" in label:
        return "backstroke"
    if "pecho" in label:
        return "breaststroke"
    if "mariposa" in label:
        return "butterfly"
    if "crol" in label or "libre" in label:
        return "freestyle"
    return None


def infer_gender_from_sheet_name(sheet_name: str) -> str | None:
    normalized = normalize_name(sheet_name)
    if "dama" in normalized or "mujer" in normalized or "female" in normalized:
        return "female"
    if "varon" in normalized or "hombre" in normalized or "male" in normalized:
        return "male"
    return None


def infer_gender_from_block(rows: list[tuple[object, ...]], header_row_index: int, sheet_name: str) -> str | None:
    sheet_gender = infer_gender_from_sheet_name(sheet_name)
    if sheet_gender:
        return sheet_gender
    for row in rows[header_row_index + 1: min(len(rows), header_row_index + 4)]:
        row_text = normalize_name(" ".join(str(value or "") for value in row))
        if "dama" in row_text or "mujer" in row_text:
            return "female"
        if "varon" in row_text or "hombre" in row_text:
            return "male"
    return None


def athlete_header_index(row: tuple[object, ...]) -> dict[str, int] | None:
    normalized = [normalize_name(str(value or "")) for value in row]
    if "nombre" not in normalized or "rut" not in normalized or not any("nacimiento" in value for value in normalized):
        return None
    header_index: dict[str, int] = {}
    for column_index, header in enumerate(normalized):
        if header == "nombre":
            header_index["first_name"] = column_index
        elif "apellido" in header and "paterno" in header:
            header_index["paternal"] = column_index
        elif "apellido" in header and "materno" in header:
            header_index["maternal"] = column_index
        elif header == "rut":
            header_index["rut"] = column_index
        elif "nacimiento" in header:
            header_index["birth_date"] = column_index
    return header_index


def event_columns_before_header(rows: list[tuple[object, ...]], header_row_index: int) -> dict[str, int]:
    event_columns: dict[str, int] = {}
    # Estos templates ponen las pruebas algunas filas arriba del header de atletas.
    for row in rows[max(0, header_row_index - 4):header_row_index]:
        for column_index, value in enumerate(row):
            stroke = stroke_from_event_label(value)
            if stroke:
                event_columns[stroke] = column_index
    return event_columns


def parse_individual_blocks_sheet(sheet_name: str, rows: list[tuple[object, ...]], competition_year: int) -> list[RelayAthlete]:
    header_rows: list[tuple[int, dict[str, int]]] = []
    for index, row in enumerate(rows):
        header_index = athlete_header_index(row)
        if header_index:
            header_rows.append((index, header_index))
    athletes: list[RelayAthlete] = []
    for block_number, (header_row_index, header_index) in enumerate(header_rows):
        gender = infer_gender_from_block(rows, header_row_index, sheet_name)
        if gender is None:
            continue
        block_end = header_rows[block_number + 1][0] if block_number + 1 < len(header_rows) else len(rows)
        event_columns = event_columns_before_header(rows, header_row_index)
        for row_number, row in enumerate(rows[header_row_index + 1:block_end], start=header_row_index + 2):
            first_name = row_cell(row, header_index.get("first_name"))
            paternal = row_cell(row, header_index.get("paternal"))
            rut_value = row_cell(row, header_index.get("rut"))
            if not first_name and not paternal and not rut_value:
                continue
            times = {
                stroke: RelayTime(ms=parse_seed_time(row_cell(row, column_index)), source="excel")
                for stroke, column_index in event_columns.items()
            }
            times = {stroke: times.get(stroke, RelayTime(ms=None, source="missing")) for stroke in STROKES}
            athlete = make_excel_athlete(
                row_number=row_number,
                first_name=first_name,
                paternal=paternal,
                maternal=row_cell(row, header_index.get("maternal")),
                rut_value=rut_value,
                birth_value=row_cell(row, header_index.get("birth_date")),
                gender_value=gender,
                times=times,
                competition_year=competition_year,
            )
            if athlete:
                athletes.append(athlete)
    return athletes


def parse_san_bernardo_template_sheet(sheet_name: str, rows: list[tuple[object, ...]], competition_year: int) -> list[RelayAthlete]:
    return parse_individual_blocks_sheet(sheet_name, rows, competition_year)


def parse_entries_workbook(file: BinaryIO | str | Path, competition_year: int = COMPETITION_YEAR) -> list[RelayAthlete]:
    workbook = load_workbook(file, data_only=True, read_only=True)
    first_sheet_rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    forms_athletes = parse_forms_style_sheet(first_sheet_rows, competition_year)
    if forms_athletes is not None:
        return forms_athletes

    athletes: list[RelayAthlete] = []
    for sheet_name in workbook.sheetnames:
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        athletes.extend(parse_san_bernardo_template_sheet(sheet_name, rows, competition_year))
    if athletes:
        return athletes
    raise ValueError("No se encontró un formato de inscripción compatible en el Excel")


def best_time_key(full_name: str, gender: str, birth_year: int | None, stroke: str) -> tuple[str, str, int | None, str]:
    return (normalize_name_match_key(full_name), gender, birth_year, stroke)


def enrich_athletes_with_db_times(athletes: list[RelayAthlete], db_best_times: dict[tuple[str, str, int | None, str], RelayTime]) -> list[RelayAthlete]:
    enriched: list[RelayAthlete] = []
    for athlete in athletes:
        times: dict[str, RelayTime] = {}
        core_athlete_id = athlete.core_athlete_id
        for stroke in STROKES:
            db_time = db_best_times.get(best_time_key(athlete.full_name, athlete.gender, athlete.birth_year, stroke))
            if db_time and db_time.ms is not None:
                times[stroke] = db_time
                core_athlete_id = db_time.athlete_core_id or core_athlete_id
            else:
                times[stroke] = RelayTime(ms=None, source="missing")
        enriched.append(replace(athlete, core_athlete_id=core_athlete_id, times=times))
    return enriched


def athlete_to_dict(athlete: RelayAthlete) -> dict[str, object]:
    return {
        "id": athlete.id,
        "core_athlete_id": athlete.core_athlete_id,
        "full_name": athlete.full_name,
        "gender": athlete.gender,
        "birth_date": athlete.birth_date.isoformat() if athlete.birth_date else None,
        "birth_year": athlete.birth_year,
        "age": athlete.age,
        "rut": athlete.rut,
        "times": {
            stroke: {
                "ms": time.ms,
                "text": time.text or format_time(time.ms),
                "source": time.source,
                "competition_name": time.competition_name,
                "competition_date": time.competition_date,
            }
            for stroke, time in athlete.times.items()
        },
    }


def analyze_athletes(athletes: list[RelayAthlete], relay_type: str = "4x50_medley_mixed") -> dict[str, object]:
    proposal, alternatives = propose_lineups(athletes, relay_type)
    used = {leg.athlete_id for lineup in proposal for leg in lineup.legs if leg.athlete_id}
    selected_type = relay_type_to_dict(relay_type)
    return {
        "competition_year": COMPETITION_YEAR,
        "relay_type": selected_type,
        "relay_types": [relay_type_to_dict(key) for key in RELAY_TYPES],
        "relay_event": selected_type["label"],
        "strokes": [{"key": stroke, "label": STROKE_LABELS[stroke]} for stroke in STROKES],
        "categories": [
            {"key": key, "label": label, "min_age_sum": minimum, "max_age_sum": maximum}
            for key, label, minimum, maximum in RELAY_CATEGORIES
        ],
        "athletes": [athlete_to_dict(athlete) for athlete in athletes],
        "proposal": [lineup_to_dict(lineup) for lineup in proposal],
        "alternatives": {key: [lineup_to_dict(lineup) for lineup in value] for key, value in alternatives.items()},
        "unassigned_athlete_ids": [athlete.id for athlete in athletes if athlete.id not in used],
    }


def roster_response(athletes: list[RelayAthlete], relay_type: str = "4x50_medley_mixed") -> dict[str, object]:
    selected_type = relay_type_to_dict(relay_type)
    return {
        "competition_year": COMPETITION_YEAR,
        "relay_type": selected_type,
        "relay_types": [relay_type_to_dict(key) for key in RELAY_TYPES],
        "relay_event": selected_type["label"],
        "strokes": [{"key": stroke, "label": STROKE_LABELS[stroke]} for stroke in STROKES],
        "categories": [
            {"key": key, "label": label, "min_age_sum": minimum, "max_age_sum": maximum}
            for key, label, minimum, maximum in RELAY_CATEGORIES
        ],
        "athletes": [athlete_to_dict(athlete) for athlete in athletes],
        "proposal": [],
        "alternatives": {key: [] for key, *_ in RELAY_CATEGORIES},
        "unassigned_athlete_ids": [athlete.id for athlete in athletes],
    }


def leg_to_dict(leg: RelayLeg) -> dict[str, object]:
    return {
        "slot_key": leg.slot_key,
        "slot_label": leg.slot_label,
        "leg_order": leg.leg_order,
        "stroke": leg.stroke,
        "stroke_label": leg.stroke_label,
        "athlete_id": leg.athlete_id,
        "athlete_name": leg.athlete_name,
        "gender": leg.gender,
        "age": leg.age,
        "time_ms": leg.time_ms,
        "time_text": leg.time_text,
        "time_source": leg.time_source,
    }


def validation_to_dict(validation: RelayValidation) -> dict[str, object]:
    return {
        "is_valid": validation.is_valid,
        "category_key": validation.category_key,
        "category_label": validation.category_label,
        "age_sum": validation.age_sum,
        "total_time_ms": validation.total_time_ms,
        "total_time_text": validation.total_time_text,
        "errors": validation.errors,
        "warnings": validation.warnings,
    }


def lineup_to_dict(lineup: RelayLineup) -> dict[str, object]:
    return {
        "id": lineup.id,
        "relay_type": lineup.relay_type,
        "category_key": lineup.category_key,
        "category_label": lineup.category_label,
        "age_sum": lineup.age_sum,
        "total_time_ms": lineup.total_time_ms,
        "total_time_text": lineup.total_time_text,
        "legs": [leg_to_dict(leg) for leg in lineup.legs],
        "validation": validation_to_dict(lineup.validation),
    }


def build_leg(slot: RelaySlot, athlete: RelayAthlete | None) -> RelayLeg:
    time = athlete.times.get(slot.stroke) if athlete else None
    return RelayLeg(
        slot_key=slot.key,
        slot_label=slot.label,
        leg_order=slot.leg_order,
        stroke=slot.stroke,
        stroke_label=slot.stroke_label,
        athlete_id=athlete.id if athlete else None,
        athlete_name=athlete.full_name if athlete else None,
        gender=athlete.gender if athlete else None,
        age=athlete.age if athlete else None,
        time_ms=time.ms if time else None,
        time_text=(time.text or format_time(time.ms)) if time else None,
        time_source=time.source if time else None,
    )


def validate_relay(legs: Iterable[RelayLeg], relay_type: str) -> RelayValidation:
    legs = list(legs)
    config = get_relay_config(relay_type)
    errors: list[str] = []
    warnings: list[str] = []
    assigned = [leg for leg in legs if leg.athlete_id]
    ids = [leg.athlete_id for leg in assigned]

    if len(assigned) != 4:
        errors.append("El relevo debe tener 4 postas asignadas.")
    if len(set(ids)) != len(ids):
        errors.append("Hay atletas repetidos dentro del relevo.")
    if any(leg.time_ms is None for leg in legs):
        errors.append("Todas las postas deben tener tiempo registrado en BD para su estilo.")
    if any(leg.age is None for leg in legs):
        errors.append("Todas las postas deben tener fecha de nacimiento válida.")

    female_count = sum(1 for leg in legs if leg.gender == "female")
    male_count = sum(1 for leg in legs if leg.gender == "male")
    gender_rule = config["gender_rule"]
    if gender_rule == "mixed_2f_2m" and (female_count != 2 or male_count != 2):
        errors.append("El relevo mixto debe tener exactamente 2 mujeres y 2 hombres.")
    if gender_rule == "women" and female_count != 4:
        errors.append("El relevo mujeres debe tener 4 mujeres.")
    if gender_rule == "men" and male_count != 4:
        errors.append("El relevo hombres debe tener 4 hombres.")

    ages = [leg.age for leg in legs]
    age_sum = sum(age for age in ages if age is not None) if all(age is not None for age in ages) else None
    category_key, category_label = relay_category(age_sum)
    if age_sum is not None and category_key is None:
        errors.append("La suma de edades no cae en ninguna categoría configurada.")

    times = [leg.time_ms for leg in legs]
    total_time_ms = sum(time for time in times if time is not None) if all(time is not None for time in times) else None
    return RelayValidation(
        is_valid=not errors,
        category_key=category_key,
        category_label=category_label,
        age_sum=age_sum,
        total_time_ms=total_time_ms,
        total_time_text=format_time(total_time_ms),
        errors=errors,
        warnings=warnings,
    )


def make_lineup(lineup_id: str, relay_type: str, athletes_by_slot: dict[str, RelayAthlete | None]) -> RelayLineup:
    legs = [build_leg(slot, athletes_by_slot.get(slot.key)) for slot in relay_slots(relay_type)]
    validation = validate_relay(legs, relay_type)
    return RelayLineup(
        id=lineup_id,
        relay_type=relay_type,
        category_key=validation.category_key,
        category_label=validation.category_label,
        age_sum=validation.age_sum,
        total_time_ms=validation.total_time_ms,
        total_time_text=validation.total_time_text,
        legs=legs,
        validation=validation,
    )


def eligible_groups(athletes: list[RelayAthlete], relay_type: str) -> Iterable[tuple[RelayAthlete, ...]]:
    config = get_relay_config(relay_type)
    women = [athlete for athlete in athletes if athlete.gender == "female" and athlete.age is not None]
    men = [athlete for athlete in athletes if athlete.gender == "male" and athlete.age is not None]
    if config["gender_rule"] == "mixed_2f_2m":
        for selected_women in combinations(women, 2):
            for selected_men in combinations(men, 2):
                yield selected_women + selected_men
    elif config["gender_rule"] == "women":
        yield from combinations(women, 4)
    elif config["gender_rule"] == "men":
        yield from combinations(men, 4)


def generate_candidate_lineups(athletes: list[RelayAthlete], relay_type: str, per_category_limit: int = 40) -> dict[str, list[RelayLineup]]:
    candidates: dict[str, list[RelayLineup]] = {key: [] for key, *_ in RELAY_CATEGORIES}
    slots = relay_slots(relay_type)
    for selected_tuple in eligible_groups(athletes, relay_type):
        selected = list(selected_tuple)
        age_sum = sum(athlete.age or 0 for athlete in selected)
        category_key, _ = relay_category(age_sum)
        if category_key is None:
            continue
        best_lineup: RelayLineup | None = None
        for ordered in permutations(selected, 4):
            by_slot = dict(zip([slot.key for slot in slots], ordered))
            if any(athlete.times.get(slot.stroke, RelayTime(None, "missing")).ms is None for slot, athlete in zip(slots, ordered)):
                continue
            lineup = make_lineup(f"candidate-{category_key}-{len(candidates[category_key])}", relay_type, by_slot)
            if not lineup.validation.is_valid:
                continue
            if best_lineup is None or (lineup.total_time_ms or 10**12) < (best_lineup.total_time_ms or 10**12):
                best_lineup = lineup
        if best_lineup:
            candidates[category_key].append(best_lineup)

    for category_key in candidates:
        unique: dict[tuple[str | None, ...], RelayLineup] = {}
        for candidate in candidates[category_key]:
            athlete_key = tuple(sorted(leg.athlete_id for leg in candidate.legs))
            current = unique.get(athlete_key)
            if current is None or (candidate.total_time_ms or 10**12) < (current.total_time_ms or 10**12):
                unique[athlete_key] = candidate
        candidates[category_key] = [
            replace(lineup, id=f"proposal-{category_key}-{index + 1}")
            for index, lineup in enumerate(sorted(unique.values(), key=lambda lineup: lineup.total_time_ms or 10**12)[:per_category_limit])
        ]
    return candidates


def propose_lineups(athletes: list[RelayAthlete], relay_type: str) -> tuple[list[RelayLineup], dict[str, list[RelayLineup]]]:
    candidates_by_category = generate_candidate_lineups(athletes, relay_type)
    category_keys = [key for key, *_ in RELAY_CATEGORIES]

    def score(lineups: list[RelayLineup]) -> tuple[int, int]:
        return (len(lineups), sum(lineup.total_time_ms or 10**12 for lineup in lineups))

    def is_better(candidate_score: tuple[int, int], current_score: tuple[int, int]) -> bool:
        return candidate_score[0] > current_score[0] or (candidate_score[0] == current_score[0] and candidate_score[1] < current_score[1])

    memo: dict[tuple[int, tuple[str, ...]], list[RelayLineup]] = {}

    def dfs(category_index: int, used_ids: frozenset[str]) -> list[RelayLineup]:
        if category_index >= len(category_keys):
            return []
        state_key = (category_index, tuple(sorted(used_ids)))
        if state_key in memo:
            return memo[state_key]
        category_key = category_keys[category_index]
        best_suffix = dfs(category_index + 1, used_ids)
        for lineup in candidates_by_category.get(category_key, []):
            ids = frozenset(leg.athlete_id for leg in lineup.legs if leg.athlete_id)
            if ids.isdisjoint(used_ids):
                candidate_suffix = [lineup, *dfs(category_index + 1, used_ids | ids)]
                if is_better(score(candidate_suffix), score(best_suffix)):
                    best_suffix = candidate_suffix
        memo[state_key] = best_suffix
        return best_suffix

    return dfs(0, frozenset()), {key: value[:5] for key, value in candidates_by_category.items()}


def relay_type_to_dict(key: str) -> dict[str, object]:
    config = get_relay_config(key)
    return {
        "key": key,
        "label": config["label"],
        "distance_m": config["distance_m"],
        "style": config["style"],
        "gender_rule": config["gender_rule"],
        "slots": [
            {"key": slot.key, "label": slot.label, "leg_order": slot.leg_order, "stroke": slot.stroke, "stroke_label": slot.stroke_label}
            for slot in relay_slots(key)
        ],
    }


def analyze_entries(file: BinaryIO | str | Path, relay_type: str = "4x50_medley_mixed", db_best_times: dict[tuple[str, str, int | None, str], RelayTime] | None = None) -> dict[str, object]:
    athletes = parse_entries_workbook(file)
    if db_best_times is not None:
        athletes = enrich_athletes_with_db_times(athletes, db_best_times)
    return analyze_athletes(athletes, relay_type)


def analyze_upload(filename: str, upload: BinaryIO, relay_type: str = "4x50_medley_mixed", db_best_times: dict[tuple[str, str, int | None, str], RelayTime] | None = None) -> dict[str, object]:
    # On Windows, NamedTemporaryFile keeps an exclusive handle that prevents
    # openpyxl from reopening the same path, so analyze uploads in memory.
    _ = filename
    return analyze_entries(BytesIO(upload.read()), relay_type=relay_type, db_best_times=db_best_times)
