from datetime import datetime
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import Workbook

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from natacion_chile.relays import (
    RelayTime,
    analyze_upload,
    compute_age,
    format_time,
    parse_entries_workbook,
    parse_seed_time,
    best_time_key,
    enrich_athletes_with_db_times,
    normalize_name_match_key,
    normalize_rut,
    roster_response,
    propose_lineups,
    relay_category,
    relay_type_to_dict,
)


def make_entries_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Control de pago",
        "Monto",
        "Marca temporal",
        "Primer Nombre",
        "Apellido Paterno",
        "Apellido Materno",
        "RUT",
        "Fecha de Nacimiento",
        "Género",
        "100 Combinado",
        "50 Pecho",
        "50 Espalda",
        "50 Mariposa",
        "50 Crol",
    ])
    ws.append([None, None, None, "Ana", "Uno", "", "1", datetime(1990, 1, 1), "FEMENINO", None, "00:40.00", "00:35.00", None, None])
    ws.append([None, None, None, "Bea", "Dos", "", "2", datetime(1991, 1, 1), "FEMENINO", None, None, None, "00:36.00", "00:31.00"])
    ws.append([None, None, None, "Carlos", "Tres", "", "3", datetime(1988, 1, 1), "MASCULINO", None, "00:34.00", "00:32.00", None, None])
    ws.append([None, None, None, "Diego", "Cuatro", "", "4", datetime(1989, 1, 1), "MASCULINO", None, None, None, "00:30.00", "00:28.00"])
    wb.save(path)




def make_san_bernardo_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Individuales DAMAS"
    ws.append([None, None, None, None, "X TORNEO MASTER SAN BERNARDO"])
    for _ in range(8):
        ws.append([])
    ws.append([None, None, None, None, None, None, None, None, "100 COMBINADO", "50 MARIPOSA", "50 ESPALDA", "50 PECHO", "50 CROL"])
    ws.append([])
    ws.append(["N?", "NOMBRE", "APELLIDO PATERNO", "APELLIDO MATERNO", "RUT", "FECHA NACIMIENTO", "EDAD"])
    ws.append(["DAMAS"])
    ws.append([1, "Lucia", "Peters", "Del Pino", "21.318.525-k", datetime(2001, 3, 14), 25, None, None, None, None, "00:48.00", None])

    ws_men = wb.create_sheet("Individuales VARONES")
    ws_men.append([None, None, None, None, "X TORNEO MASTER SAN BERNARDO"])
    for _ in range(8):
        ws_men.append([])
    ws_men.append([None, None, None, None, None, None, None, None, "100 COMBINADO", "50 MARIPOSA", "50 ESPALDA", "50 PECHO", "50 CROL"])
    ws_men.append([])
    ws_men.append(["N?", "NOMBRE", "APELLIDO PATERNO", "APELLIDO MATERNO", "RUT", "FECHA NACIMIENTO", "EDAD"])
    ws_men.append(["VARONES"])
    ws_men.append([1, "Ricardo", "Cabas", "Galindo", "28.026.125-4", datetime(2000, 5, 17), 26, None, None, None, "00:33.00", None, "00:27.50"])

    wb.create_sheet("RELEVOS 4x50 Mixto Combinado")
    wb.save(path)



def make_single_sheet_block_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "INDIVIDUALES"
    ws.append([None, None, None, None, "VI COPA SANTIAGO DEPORTE"])
    for _ in range(8):
        ws.append([])
    ws.append([None, None, None, None, None, None, None, None, "200 COMBINADO", "50 ESPALDA", "100 MARIPOSA", "100 PECHO", "50 CROL"])
    ws.append([])
    ws.append(["N?", "NOMBRE", "APELLIDO PATERNO", "APELLIDO MATERNO", "RUT", "FECHA NACIMIENTO", "EDAD"])
    ws.append(["DAMAS INDIVIDUALES", None, None, None, None, None, None, None, "Prueba N?01", "Prueba N?02", "Prueba N?03", "Prueba N?04", "Prueba N?05"])
    ws.append([1, "Ana", "Uno", "Dos", "11.111.111-1", datetime(1990, 1, 1), 36, None, None, "00:40.50", None, None, "00:32.00"])
    ws.append([])
    ws.append(["Todos los datos son obligatorios"])
    ws.append([None, None, None, None, None, None, None, None, "200 COMBINADO", "50 ESPALDA", "100 MARIPOSA", "100 PECHO", "50 CROL"])
    ws.append([])
    ws.append(["N?", "NOMBRE", "APELLIDO PATERNO", "APELLIDO MATERNO", "RUT", "FECHA NACIMIENTO", "EDAD"])
    ws.append(["VARONES INDIVIDUALES", None, None, None, None, None, None, None, "Prueba N?01", "Prueba N?02", "Prueba N?03", "Prueba N?04", "Prueba N?05"])
    ws.append([1, "Bruno", "Tres", "Cuatro", "22.222.222-2", datetime(1988, 1, 1), 38, None, None, None, None, "01:20.00", "00:29.75"])
    wb.save(path)
def test_parse_seed_time_accepts_minutes_seconds_and_hundredths():
    assert parse_seed_time("01:19.44") == 79_440
    assert parse_seed_time("30.96") == 30_960
    assert format_time(79_440) == "01:19.44"


def test_compute_age_uses_competition_year():
    assert compute_age(datetime(1993, 11, 30), 2026) == 33


def test_relay_category_by_age_sum():
    assert relay_category(96) == ("premaster", "Premaster")
    assert relay_category(120) == ("120-159", "120 - 159")
    assert relay_category(99) == (None, None)


def test_parse_entries_workbook_reads_athletes_and_times(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)

    athletes = parse_entries_workbook(path)

    assert len(athletes) == 4
    assert athletes[0].gender == "female"
    assert athletes[0].times["backstroke"].ms == 35_000
    assert athletes[1].times["freestyle"].ms == 31_000


def test_parse_entries_workbook_reads_san_bernardo_template(tmp_path):
    path = tmp_path / "san-bernardo.xlsx"
    make_san_bernardo_workbook(path)

    athletes = parse_entries_workbook(path)

    assert len(athletes) == 2
    assert athletes[0].full_name == "Lucia Peters Del Pino"
    assert athletes[0].gender == "female"
    assert athletes[0].rut == "21318525K"
    assert athletes[0].times["breaststroke"].ms == 48_000
    assert athletes[1].gender == "male"
    assert athletes[1].times["backstroke"].ms == 33_000
    assert athletes[1].times["freestyle"].ms == 27_500


def test_parse_entries_workbook_reads_single_sheet_gender_blocks(tmp_path):
    path = tmp_path / "copa-santiago.xlsx"
    make_single_sheet_block_workbook(path)

    athletes = parse_entries_workbook(path)

    assert len(athletes) == 2
    assert athletes[0].full_name == "Ana Uno Dos"
    assert athletes[0].gender == "female"
    assert athletes[0].times["backstroke"].ms == 40_500
    assert athletes[0].times["freestyle"].ms == 32_000
    assert athletes[1].full_name == "Bruno Tres Cuatro"
    assert athletes[1].gender == "male"
    assert athletes[1].times["breaststroke"].ms is None
    assert athletes[1].times["freestyle"].ms == 29_750


def test_analyze_upload_reads_excel_from_memory_without_temp_file_lock(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)

    result = analyze_upload("entries.xlsx", BytesIO(path.read_bytes()))

    assert result["competition_year"] == 2026
    assert len(result["athletes"]) == 4
    assert len(result["proposal"]) == 1


def test_propose_lineups_returns_valid_mixed_relay_without_reuse(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)
    athletes = parse_entries_workbook(path)

    proposal, alternatives = propose_lineups(athletes, "4x50_medley_mixed")

    assert len(proposal) == 1
    lineup = proposal[0]
    assert lineup.validation.is_valid
    assert lineup.validation.category_key == "120-159"
    assert lineup.total_time_ms == 130_000
    assert {leg.gender for leg in lineup.legs} == {"female", "male"}
    assert len({leg.athlete_id for leg in lineup.legs}) == 4
    assert alternatives["120-159"]


def test_db_times_replace_excel_times_for_relay_analysis(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)
    athletes = parse_entries_workbook(path)
    db_times = {
        best_time_key("Ana Uno", "female", 1990, "backstroke"): RelayTime(ms=31_000, source="db", athlete_core_id=10),
        best_time_key("Ana Uno", "female", 1990, "breaststroke"): RelayTime(ms=39_000, source="db", athlete_core_id=10),
    }

    enriched = enrich_athletes_with_db_times(athletes, db_times)

    assert enriched[0].core_athlete_id == 10
    assert enriched[0].times["backstroke"].ms == 31_000
    assert enriched[0].times["backstroke"].source == "db"
    assert enriched[0].times["freestyle"].ms is None
    assert enriched[0].times["freestyle"].source == "missing"




def test_relay_types_include_generated_50_and_100_variants():
    keys = {relay_type_to_dict(key)["key"] for key in [
        "4x50_medley_mixed",
        "4x50_medley_women",
        "4x50_medley_men",
        "4x50_freestyle_mixed",
        "4x50_freestyle_women",
        "4x50_freestyle_men",
        "4x100_medley_mixed",
        "4x100_medley_women",
        "4x100_medley_men",
        "4x100_freestyle_mixed",
        "4x100_freestyle_women",
        "4x100_freestyle_men",
    ]}

    assert len(keys) == 12
    relay_type = relay_type_to_dict("4x100_medley_mixed")
    assert relay_type["distance_m"] == 100
    assert relay_type["style"] == "medley"
    assert [slot["stroke"] for slot in relay_type["slots"]] == ["backstroke", "breaststroke", "butterfly", "freestyle"]

def test_freestyle_relay_type_uses_four_freestyle_slots(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)
    athletes = parse_entries_workbook(path)
    db_times = {
        best_time_key(athlete.full_name, athlete.gender, athlete.birth_year, "freestyle"): RelayTime(ms=30_000 + index, source="db")
        for index, athlete in enumerate(athletes)
    }
    athletes = enrich_athletes_with_db_times(athletes, db_times)

    proposal, _ = propose_lineups(athletes, "4x50_freestyle_mixed")

    assert len(proposal) == 1
    assert [leg.stroke for leg in proposal[0].legs] == ["freestyle", "freestyle", "freestyle", "freestyle"]
    assert {leg.time_source for leg in proposal[0].legs} == {"db"}
    assert proposal[0].validation.is_valid


def test_name_match_key_tolerates_canonical_and_natural_order():
    assert normalize_name_match_key("ALEXIS SAYAGO MORENO") == normalize_name_match_key("SAYAGO MORENO, ALEXIS")


def test_normalize_rut_removes_punctuation_and_uppercases_k():
    assert normalize_rut("12.345.678-k") == "12345678K"
    assert normalize_rut(" 25.436.854-7 ") == "254368547"


def test_roster_response_does_not_generate_proposals(tmp_path):
    path = tmp_path / "entries.xlsx"
    make_entries_workbook(path)
    athletes = parse_entries_workbook(path)

    response = roster_response(athletes, "4x50_medley_mixed")

    assert len(response["athletes"]) == 4
    assert response["proposal"] == []
    assert all(value == [] for value in response["alternatives"].values())
